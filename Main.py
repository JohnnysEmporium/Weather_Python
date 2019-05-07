# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'window_final.ui'
#
# Created by: PyQt5 UI code generator 5.9.2
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.Qt import QPixmap, QFont, QImage, QSize, QPalette, QBrush, QColor
from PIL import Image
from scipy.ndimage import imread
from PIL.ImageQt import ImageQt
import pyqtgraph as pg
import openpyxl as xl
import requests
import json
import time
import threading


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(749, 733)
        MainWindow.setStyleSheet("")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.gridFor2 = QtWidgets.QGridLayout()
        self.gridFor2.setObjectName("gridFor2")
        self.scrollArea_2 = QtWidgets.QScrollArea(self.centralwidget)
        self.scrollArea_2.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.scrollArea_2.setWidgetResizable(True)
        self.scrollArea_2.setObjectName("scrollArea_2")
        self.hForecastSA2 = QtWidgets.QWidget()
        self.hForecastSA2.setGeometry(QtCore.QRect(0, 0, 727, 276))
        self.hForecastSA2.setObjectName("hForecastSA2")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.hForecastSA2)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.scrollArea_2_layout = QtWidgets.QGridLayout()
        self.scrollArea_2_layout.setObjectName("scrollArea_2_layout")
        self.gridLayout_2.addLayout(self.scrollArea_2_layout, 0, 0, 1, 1)
        self.scrollArea_2.setWidget(self.hForecastSA2)
        self.gridFor2.addWidget(self.scrollArea_2, 1, 0, 1, 1)
        self.scrollArea_1 = QtWidgets.QScrollArea(self.centralwidget)
        self.scrollArea_1.setAutoFillBackground(False)
        self.scrollArea_1.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.scrollArea_1.setWidgetResizable(True)
        self.scrollArea_1.setObjectName("scrollArea_1")
        self.hForecastSA1 = QtWidgets.QWidget()
        self.hForecastSA1.setGeometry(QtCore.QRect(0, 0, 727, 277))
        self.hForecastSA1.setObjectName("hForecastSA1")
        self.gridLayout_4 = QtWidgets.QGridLayout(self.hForecastSA1)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.scrollArea_1_layout = QtWidgets.QGridLayout()
        self.scrollArea_1_layout.setObjectName("scrollArea_1_layout")
        self.gridLayout_4.addLayout(self.scrollArea_1_layout, 0, 0, 1, 1)
        self.scrollArea_1.setWidget(self.hForecastSA1)
        self.gridFor2.addWidget(self.scrollArea_1, 0, 0, 1, 1)
        self.gridLayout.addLayout(self.gridFor2, 3, 0, 1, 1)
        self.gridLayout_3 = QtWidgets.QGridLayout()
        self.gridLayout_3.setObjectName("gridLayout_3")
        spacerItem = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_3.addItem(spacerItem, 0, 2, 4, 1)
        self.humidityimg = QtWidgets.QLabel(self.centralwidget)
        self.humidityimg.setAlignment(QtCore.Qt.AlignCenter)
        self.humidityimg.setObjectName("humidityimg")
        self.gridLayout_3.addWidget(self.humidityimg, 2, 0, 1, 1)
        self.cloudimg = QtWidgets.QLabel(self.centralwidget)
        self.cloudimg.setAlignment(QtCore.Qt.AlignCenter)
        self.cloudimg.setObjectName("cloudimg")
        self.gridLayout_3.addWidget(self.cloudimg, 1, 0, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_3.addItem(spacerItem1, 0, 4, 4, 1)
        self.cloud = QtWidgets.QLabel(self.centralwidget)
        self.cloud.setAlignment(QtCore.Qt.AlignCenter)
        self.cloud.setObjectName("cloud")
        self.gridLayout_3.addWidget(self.cloud, 1, 3, 1, 1)
        self.pressure = QtWidgets.QLabel(self.centralwidget)
        self.pressure.setAlignment(QtCore.Qt.AlignCenter)
        self.pressure.setObjectName("pressure")
        self.gridLayout_3.addWidget(self.pressure, 3, 3, 1, 1)
        self.humidity = QtWidgets.QLabel(self.centralwidget)
        self.humidity.setAlignment(QtCore.Qt.AlignCenter)
        self.humidity.setObjectName("humidity")
        self.gridLayout_3.addWidget(self.humidity, 2, 3, 1, 1)
        self.pressureimg = QtWidgets.QLabel(self.centralwidget)
        self.pressureimg.setAlignment(QtCore.Qt.AlignCenter)
        self.pressureimg.setObjectName("pressureimg")
        self.gridLayout_3.addWidget(self.pressureimg, 3, 0, 1, 1)
        self.temp = QtWidgets.QLabel(self.centralwidget)
        self.temp.setAlignment(QtCore.Qt.AlignCenter)
        self.temp.setObjectName("temp")
        self.gridLayout_3.addWidget(self.temp, 0, 3, 1, 1)
        self.tempimg = QtWidgets.QLabel(self.centralwidget)
        self.tempimg.setAlignment(QtCore.Qt.AlignCenter)
        self.tempimg.setObjectName("tempimg")
        self.gridLayout_3.addWidget(self.tempimg, 0, 0, 1, 1)
        self.windimg = QtWidgets.QLabel(self.centralwidget)
        self.windimg.setAlignment(QtCore.Qt.AlignCenter)
        self.windimg.setObjectName("windimg")
        self.gridLayout_3.addWidget(self.windimg, 0, 5, 1, 1)
        spacerItem2 = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        self.gridLayout_3.addItem(spacerItem2, 0, 6, 4, 1)
        self.winddirimg = QtWidgets.QLabel(self.centralwidget)
        self.winddirimg.setAlignment(QtCore.Qt.AlignCenter)
        self.winddirimg.setObjectName("winddirimg")
        self.gridLayout_3.addWidget(self.winddirimg, 1, 5, 1, 1)
        self.rainimg = QtWidgets.QLabel(self.centralwidget)
        self.rainimg.setAlignment(QtCore.Qt.AlignCenter)
        self.rainimg.setObjectName("rainimg")
        self.gridLayout_3.addWidget(self.rainimg, 2, 5, 1, 1)
        self.wind = QtWidgets.QLabel(self.centralwidget)
        self.wind.setAlignment(QtCore.Qt.AlignCenter)
        self.wind.setObjectName("wind")
        self.gridLayout_3.addWidget(self.wind, 0, 7, 1, 1)
        self.winddir = QtWidgets.QLabel(self.centralwidget)
        self.winddir.setAlignment(QtCore.Qt.AlignCenter)
        self.winddir.setObjectName("winddir")
        self.gridLayout_3.addWidget(self.winddir, 1, 7, 1, 1)
        self.rain = QtWidgets.QLabel(self.centralwidget)
        self.rain.setAlignment(QtCore.Qt.AlignCenter)
        self.rain.setObjectName("rain")
        self.gridLayout_3.addWidget(self.rain, 2, 7, 1, 1)
        self.gridLayout.addLayout(self.gridLayout_3, 1, 0, 1, 1)
        self.search_grid = QtWidgets.QGridLayout()
        self.search_grid.setObjectName("search_grid")
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setObjectName("lineEdit")
        self.search_grid.addWidget(self.lineEdit, 0, 0, 1, 1)
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setObjectName("pushButton")
        self.search_grid.addWidget(self.pushButton, 0, 1, 1, 1)
        self.gridLayout.addLayout(self.search_grid, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 749, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.humidityimg.setText(_translate("MainWindow", "humidityimg"))
        self.cloudimg.setText(_translate("MainWindow", "cloudimg"))
        self.cloud.setText(_translate("MainWindow", "cloud"))
        self.pressure.setText(_translate("MainWindow", "pressure"))
        self.humidity.setText(_translate("MainWindow", "humidity"))
        self.pressureimg.setText(_translate("MainWindow", "pressureimg"))
        self.temp.setText(_translate("MainWindow", "temp"))
        self.tempimg.setText(_translate("MainWindow", "tempimg"))
        self.windimg.setText(_translate("MainWindow", "windimg"))
        self.winddirimg.setText(_translate("MainWindow", "windirimg"))
        self.rainimg.setText(_translate("MainWindow", "rainimg"))
        self.wind.setText(_translate("MainWindow", "wind"))
        self.winddir.setText(_translate("MainWindow", "winddir"))
        self.rain.setText(_translate("MainWindow", "rain"))
        self.pushButton.setText(_translate("MainWindow", "Get Forecast"))

    def set_misc(self):
        
        self.retranslateUi(MainWindow)
        self.pushButton.clicked.connect(self.refresh)
        self.pushButton.setShortcut("Return")
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        oImage = QImage("icons\\background.png")
        sImage = oImage.scaled(QSize(749,733))
        palette = QPalette()
        palette.setBrush(10, QBrush(oImage))                     
        MainWindow.setPalette(palette)
        
        color = QColor(255, 0, 0, 0)
        palette = QPalette()
        palette.setBrush(10, QBrush(color))
        self.hForecastSA1.setPalette(palette)
        self.hForecastSA2.setPalette(palette)
         
        self.scrollArea_1.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.scrollArea_1.setFrameShape(QtGui.QFrame.NoFrame)
        self.scrollArea_2.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.scrollArea_2.setFrameShape(QtGui.QFrame.NoFrame)
         
        sAreaStyleSheet = """
            
            QScrollArea{
                background-color:transparent;
                }
                
            """
        
        sBarStyleSheet = """
        
            QScrollBar:horizontal{
                background-color:transparent;
                }
        
        """
        
        lineEditStyleSheet = """
        
            *{
                background-color: rgba(0, 0, 0, 50);
                color: white;
            }
        
        """
        
        pushButtonStyleSheet = """
        
            *{
                background-color: rgba(0, 0, 0, 50);
                color: white;
            }
        
        """
        
        self.scrollArea_1.setStyleSheet(sAreaStyleSheet)
        self.scrollArea_1.horizontalScrollBar().setStyleSheet(sBarStyleSheet)
        self.scrollArea_2.setStyleSheet(sAreaStyleSheet)
        self.scrollArea_2.horizontalScrollBar().setStyleSheet(sBarStyleSheet)

        self.lineEdit.setStyleSheet(lineEditStyleSheet)
        self.pushButton.setStyleSheet(pushButtonStyleSheet)


    def get_cities(self):
        self.cities = {}
        wb = xl.load_workbook('worldcities.xlsx')
        wb = wb.get_sheet_by_name('Sheet1')
        for i in range(2, wb.max_row + 1):
            self.cities[wb.cell(row = i, column = 1).value + ', ' + wb.cell(row = i, column = 5).value] = [wb.cell(row = i, column = 3).value, wb.cell(row = i, column = 4).value]
            

        
    def autocomplete(self):
        completer = QtWidgets.QCompleter()
        completer.setMaxVisibleItems(10)
        self.lineEdit.setCompleter(completer)
         
        model = QtCore.QStringListModel()
        model.setStringList(self.cities)
        completer.setModel(model)
      
    def get_data(self):
        try:
            city = self.cities[self.lineEdit.text()]
            lon = str(city[0])
            lat = str(city[1])
            self.response = requests.get('https://api.darksky.net/forecast/YOUR API KEY/' + lon + ',' + lat + '?units=si')
        except:
            self.response = requests.get('https://api.darksky.net/forecast/YOUR API KEY/50.0646,19.9450?units=si')
            
    def change_font(self):
        lab = MainWindow.findChildren(QtWidgets.QLabel)
        for i in lab[-14:]:
            i.setFont(QtGui.QFont('SansSerif', 14))
        
    def winddir_rotation(self, x):
        im = Image.open('icons\\winddir.png')
        rot = im.rotate(-1 * self.winddirs[x] - 180, resample = Image.BICUBIC, expand = True)
        rot = rot.crop((0, 0, 128, 128))
        rot = ImageQt(rot)
        rot = QPixmap.fromImage(rot)
        rot = rot.scaled(64, 64)
        self.winddirimg.setPixmap(rot)
    
    def icon_handler(self, x):
        self.setimg(self.tempimg, 'temp')
        self.setimg(self.humidityimg, 'humidity')
        self.setimg(self.pressureimg, 'pressure')
        self.setimg(self.windimg, 'wind')
        self.winddir_rotation(x)
        self.setimg(self.rainimg, 'rain')
        self.setimg(self.cloudimg, self.cloud_icon_handler(self.clouds, x))

#    0 FOR CURRENT, 1 FOR HOURLY, 2 FOR DAILY 
    def value_handler(self, x):
        self.values()
        self.temp.setText('Temperature\n' + str(round(self.temps[x], 1)) + ' °C')
        self.cloud.setText('Cloud Cover\n' + str(round(self.clouds[x] * 100)) + ' %')
        self.humidity.setText('Humidity\n' + str(round(self.humiditys[x] * 100)) + ' %')
        self.pressure.setText('Pressure\n' + str(round(self.pressures[x],1)) + ' hPa')
        self.wind.setText('Wind Speed\n' + str(round((self.winds[x] * 3600) / 1000)) + ' km/h')
        self.rain.setText('Rain Probability\n' + str(round(self.rainchances[x] * 100, 1)) + ' %')
        if 337.5 <= self.winddirs[x] < 22.5:
            self.winddir.setText('Wind Direction\n' + str('N'))
        if 22.5 <= self.winddirs[x] < 67.5:
            self.winddir.setText('Wind Direction\n' + str('NE'))
        if 67.5 <= self.winddirs[x] < 112.5:
            self.winddir.setText('Wind Direction\n' + str('E'))
        if 112.5 <= self.winddirs[x] < 157.5:
            self.winddir.setText('Wind Direction\n' + str('SE'))
        if 157.5 <= self.winddirs[x] < 202.5:
            self.winddir.setText('Wind Direction\n' + str('S'))
        if 202.5 <= self.winddirs[x] < 247.5:
            self.winddir.setText('Wind Direction\n' + str('SW'))
        if 247 <= self.winddirs[x] < 292.5:
            self.winddir.setText('Wind Direction\n' + str('W'))
        if 292.5 <= self.winddirs[x] < 337.5:
            self.winddir.setText('Wind Direction\n' + str('NW'))
            
    def values(self):
        self.htime = []
        self.dtime = []
        self.temps = []
        self.htemps = []
        self.dtemps = []
        self.clouds = []
        self.hclouds = []
        self.dclouds = []
        self.humiditys = []
        self.pressures = []
        self.winds = []
        self.winddirs = []
        self.rainchances = []
        self.rainintensitys = []
#         self.get_data()
        self.vals = self.response.json()
        self.curvals = self.vals['currently']
        self.hvals = self.vals['hourly']['data'][0]
        self.dayvals = self.vals['daily']['data'][0]
        self.summary = self.vals['daily']['summary']
        
#         CURRENT DATA
        self.temps.append(self.curvals['temperature'])
        self.clouds.append(self.curvals['cloudCover'])
        self.humiditys.append(self.curvals['humidity'])
        self.pressures.append(self.curvals['pressure'])
        self.winds.append(self.curvals['windSpeed'])
        self.winddirs.append(round(self.curvals['windBearing'], 1))
        self.rainchances.append(self.curvals['precipProbability'])
        self.rainintensitys.append(self.curvals['precipIntensity'])

#         HOURLY DATA
        for i in range(len(self.vals['hourly']['data'])):
            self.htime.append(self.vals['hourly']['data'][i]['time'])
            self.htemps.append(self.vals['hourly']['data'][i]['temperature'])
            self.hclouds.append(self.vals['hourly']['data'][i]['cloudCover'],) 
        self.htemps.append(self.hvals['temperature'])
        
        self.humiditys.append(self.hvals['humidity'])
        self.pressures.append(self.hvals['pressure'])
        self.winds.append(self.hvals['windSpeed'])
        self.winddirs.append(round(self.curvals['windBearing'], 1))
        self.rainchances.append(self.hvals['precipProbability'])
        self.rainintensitys.append(self.hvals['precipIntensity'])
        
#         DAILY DATA
        for i in range(len(self.vals['daily']['data'])):
            self.dtime.append(self.vals['daily']['data'][i]['time'])
            self.dtemps.append(self.vals['daily']['data'][i]['temperatureMax'])
            self.dclouds.append(self.vals['daily']['data'][i]['cloudCover'],) 
        self.clouds.append(self.dayvals['cloudCover'])
        self.humiditys.append(self.dayvals['humidity'])
        self.pressures.append(self.dayvals['pressure'])
        self.winds.append(self.dayvals['windSpeed'])
        self.winddirs.append(round(self.curvals['windBearing'], 1))
        self.rainchances.append(self.dayvals['precipProbability'])
        self.rainintensitys.append(self.dayvals['precipIntensity'])
        
    
    def hForecast(self):
        self.hForecastLabels = []

        for i in range(len(self.htemps) - 1):
            
            t = time.localtime(self.htime[i]).tm_hour
            
            self.hForecastLabels.append([
                QtWidgets.QLabel(self.centralwidget, text = str(t) + ':00'), 
                QtWidgets.QLabel(self.centralwidget),
                QtWidgets.QLabel(self.centralwidget, text = str(round(self.htemps[i])) + ' °C')
                ])
            
            self.setimg(self.hForecastLabels[i][1], self.cloud_icon_handler(self.hclouds, i))
            
            self.scrollArea_1_layout.addWidget(self.hForecastLabels[i][0], 0, i)
            self.scrollArea_1_layout.addWidget(self.hForecastLabels[i][1], 1, i)
            self.scrollArea_1_layout.addWidget(self.hForecastLabels[i][2], 2, i)
            
            for j in range(3):
                self.hForecastLabels[i][j].setAlignment(QtCore.Qt.AlignCenter)

    def dForecast(self):
        self.dForecastLabels = []
        
        for i in range(len(self.dtemps) - 1):
            
            t = time.localtime(self.htime[i]).tm_hour
            
            self.dForecastLabels.append([
                QtWidgets.QLabel(self.centralwidget, text = self.weekday_handler()[i]), 
                QtWidgets.QLabel(self.centralwidget),
                QtWidgets.QLabel(self.centralwidget, text = str(round(self.dtemps[i])) + ' °C')
                ])
            
            self.setimg(self.dForecastLabels[i][1], self.cloud_icon_handler(self.dclouds, i))
            
            self.scrollArea_2_layout.addWidget(self.dForecastLabels[i][0], 0, i)
            self.scrollArea_2_layout.addWidget(self.dForecastLabels[i][1], 1, i)
            self.scrollArea_2_layout.addWidget(self.dForecastLabels[i][2], 2, i)
            
            for j in range(3):
                self.dForecastLabels[i][j].setAlignment(QtCore.Qt.AlignCenter)

    def weekday_handler(self):
        var = []
        for i in range(len(self.dtime)):
            t = time.localtime(self.dtime[i]).tm_wday
            if t == 0:
                var.append('Monday')
            if t == 0:
                var.append('Tuesday')
            if t == 0:
                var.append('Wednesday')
            if t == 0:
                var.append('Thursday')
            if t == 0:
                var.append('Friday')
            if t == 0:
                var.append('Saturday')
            if t == 0:
                var.append('Sunday')
        return var
            
                
    def setimg(self, var, img):
        im = "icons\\" + img + ".png"
        im = QPixmap(im)
        im = im.scaled(64, 64)
        var.setPixmap(QPixmap(im))
        
    def cloud_icon_handler(self, var, x):
        if  var[x] < .4:
            return 'sun'
        if .4 <= var[x] <= .7:
            return 'cloudy_sun'
        if var[x] > .7:
            return 'cloud'
        
    def rm_widgets(self):
        d = self.dForecastLabels
        h = self.hForecastLabels
        for i in range(len(d)):
            for j in range(len(d[i])):
                d[i][j].hide()
        for i in range(len(h)):
            for j in range(len(h[i])):
                h[i][j].hide()
        
    def refresh(self):
#         while True:
            self.get_data()
            self.value_handler(0)
            try:
                self.rm_widgets()
            except:
                pass
            self.icon_handler(0)
            self.hForecast()
            self.dForecast()
#             time.sleep(10)
        
        
if __name__ == "__main__":
    import sys
    ui = Ui_MainWindow()
    thread0 = threading.Thread(target = ui.get_data)
    thread0.start()
    thread1 = threading.Thread(target = ui.get_cities)
    thread1.start()
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    ui.set_misc()
    ui.refresh()
    ui.change_font()
    thread1.join()
    thread2 = threading.Thread(target = ui.autocomplete())
    thread2.start()
    sys.exit(app.exec_())

